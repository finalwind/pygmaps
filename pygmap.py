
import openpyxl
import googlemaps
#get origin and destination locations



def cleanWB(file_path):
    gmaps = googlemaps.Client(key ='INSERT API KEY HERE')
    destination = list()
    wb = openpyxl.load_workbook(filename=file_path)
    ws = wb.get_sheet_by_name('Sheet1')
    for i in range(ws.max_row):
        cellValueLocation = ws.cell(row=i+2,column=1).value
        destination.append(cellValueLocation)

    #remove duplicates from destination list
    unique_location = list(set(destination))
    return unique_location

def getDistance(origin, destination):
    #Google distance matrix API key
    gmaps = googlemaps.Client(key ='INSERT API KEY HERE')

    distance = gmaps.distance_matrix(origin, destination)
    distance_status = distance['rows'][0]['elements'][0]['status']
    if distance_status != 'ZERO_RESULTS':
        jDistance = distance['rows'][0]['elements'][0]
        distance_location = jDistance['distance']['value']
    else:
        distance_location = 0

    return distance_location

def getGeocode(loc):
    gmaps = googlemaps.Client(key ='INSERT API KEY HERE')

    gcode = gmaps.geocode(loc)
    geocode_result = ()
    geocode_result = (gcode[0]['geometry']['bounds']['southwest']['lat'],gcode[0]['geometry']['bounds']['southwest']['lng'])
    return geocode_result

def getPlace(loc):
    gmaps = googlemaps.Client(key ='INSERT API KEY HERE')


    gplace = gmaps.places(loc)
    gplace_result = (gplace['results'][0]['formatted_address'])

    return gplace_result
